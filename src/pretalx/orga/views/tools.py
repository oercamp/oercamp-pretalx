import collections
import datetime as dt
import json
from contextlib import suppress
import openpyxl

import dateutil.parser
from csp.decorators import csp_update
from django.conf import settings
from django.contrib import messages
from django.db.models.deletion import ProtectedError
from django.http import FileResponse, JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.utils.decorators import method_decorator
from django.utils.functional import cached_property
from django.utils.timezone import now
from django.utils.translation import gettext_lazy as _
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, TemplateView, UpdateView, View
from django_context_decorator import context
from i18nfield.strings import LazyI18nString
from i18nfield.utils import I18nJSONEncoder

from pretalx.agenda.management.commands.export_schedule_html import get_export_zip_path
from pretalx.agenda.tasks import export_schedule_html
from pretalx.agenda.views.utils import get_schedule_exporters
from pretalx.common.language import get_current_language_information
from pretalx.common.text.path import safe_filename
from pretalx.common.text.phrases import phrases
from pretalx.common.views import CreateOrUpdateView, OrderModelView
from pretalx.common.views.mixins import (
    ActionFromUrl,
    EventPermissionRequired,
    PermissionRequired,
)
from pretalx.orga.forms.tools import (
    SurveyMergerForm,
)
from pretalx.schedule.forms import QuickScheduleForm, RoomForm
from pretalx.schedule.models import Availability, Room, TalkSlot, Schedule
from pretalx.schedule.utils import guess_schedule_version

import requests
import logging


class SurveyMergerView(EventPermissionRequired, FormView):
    template_name = "orga/tools/survey_merger.html"
    permission_required = "orga.change_settings"
    form_class = SurveyMergerForm

    def form_valid(self, form):
        uploaded_file = form.cleaned_data.get("excel_file")
        if not uploaded_file:
            return self.form_invalid(form)  # Re-render the form with an error
        # Process the file
        # Example: Read and process the file in memory
        try:
            # Add your logic to process the file here
            # Example: log file size
            messages.error(self.request, f"Uploaded file size: {uploaded_file.size} bytes")
        except Exception as e:
            form.add_error("excel_file", f"Error processing file: {e}")
            return self.form_invalid(form)


        if not self.request.event.is_pretix_api_configured:
            form.add_error("excel_file", f"Error : Pretix API is not configured.")
            return self.form_invalid(form)

        modified_file = self.process_excel_file(uploaded_file)

        # Return the modified file as a downloadable response
        response = HttpResponse(
            modified_file,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="modified_file.xlsx"'
        return response

    def get_success_url(self):
        return self.request.event.orga_urls.tools_survey_merger

    def get_form_kwargs(self):
        result = super().get_form_kwargs()
        result["event"] = self.request.event
        return result


    def get_context_data(self, **kwargs):
        result = super().get_context_data(**kwargs)
        return result


    ###
    # This Function takes the uploaded excel file, then
    # 1. requests all orders from pretix.
    # 2. connects orders and excel by token
    # 3. add headers for the questions (question_identifier is used) to the result excel
    # 4. fills the columns with answer data
    # 5. removes the token column (TODO)
    ###
    def process_excel_file(self, uploaded_file):
        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active  # or specify sheet by name: workbook['Sheet1']

        qid_whitelist = self.request.event.survey_merger_pretix_qid_whitelist
        qid_whitelist_array = [qid.strip() for qid in qid_whitelist.split(",")] if qid_whitelist else []

        logging.info(qid_whitelist_array)

        api_answer_data = self.request_answer_data() if qid_whitelist_array else []

        # Create a dictionary to map tokens to their answers and question identifiers
        token_data_map = {}
        for entry in api_answer_data:
            token = entry['token']
            question_identifier = entry['question_identifier']
            answer = entry['answer']

            if token not in token_data_map:
                token_data_map[token] = {}
            token_data_map[token][question_identifier] = answer

        # Identify the 'token' column dynamically from the header row (assuming headers are in the first row)
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
        token_column_index = headers.get("token")
        if not token_column_index:
            logging.error("Column with header 'token' not found.")
            return

        # Filter question identifiers based on the whitelist
        qid_whitelist_set = set(qid_whitelist_array)  # Convert whitelist to a set for faster lookups

        # Add headers for new columns dynamically if not already present and in the whitelist
        for entry in api_answer_data:
            question_identifier = entry['question_identifier']
            if question_identifier in qid_whitelist_set and question_identifier not in headers:
                new_col_index = len(headers) + 1
                headers[question_identifier] = new_col_index
                sheet.cell(row=1, column=new_col_index, value=question_identifier)

        # Iterate over each row, starting from the second row
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            # Get the token value from the dynamically determined token column
            token_cell = row[token_column_index - 1]  # Adjust for 0-based indexing
            token_value = token_cell.value.strip() if token_cell.value else None

            # Check if the token exists in the data map
            if token_value in token_data_map:
                for question_identifier, answer in token_data_map[token_value].items():
                    # Only process question identifiers in the whitelist
                    if question_identifier in qid_whitelist_set:
                        # Find the correct column for the question identifier
                        col_index = headers[question_identifier]
                        # Insert the answer into the appropriate cell
                        sheet.cell(row=row_index, column=col_index, value=answer)


        # Save the modified workbook to a byte stream (instead of saving to disk)
        from io import BytesIO
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)  # Move back to the start of the byte stream

        return file_stream.read()

    def request_answer_data(self):

        url = f"https://{self.request.event.pretix_api_domain}/api/v1/organizers/{self.request.event.pretix_api_organisator_slug}/events/{self.request.event.pretix_api_event_slug}/orders/"
        headers = {
            'Authorization': f"Token {self.request.event.pretix_api_key}"
        }

        result_data_list = []

        while url:  # Continue looping as long as there's a next URL
            response = requests.get(url, headers=headers)

            if response.status_code != 200:
                return {'error': 'Failed to fetch data'}

            data = response.json()  # Parse the JSON response

            # Iterate through each result in the results list
            for result in data['results']:
                # Check if the status is 'p' (= paid)
                if result['status'] == 'p':
                    # Iterate through each position in the positions list
                    for position in result['positions']:

                        answer_data = []

                        # Second loop: get all answers
                        for answer in position.get('answers', []):
                            answer_data.append(
                                {
                                    'token': result.get('secret_token'),
                                    'question_identifier': answer.get('question_identifier'),
                                    'answer': answer.get('answer')
                                }
                            )

                        result_data_list.extend(answer_data)

            # Get the URL for the next page, if any
            url = data.get('next')

        return result_data_list
